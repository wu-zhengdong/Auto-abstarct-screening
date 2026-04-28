import csv
from datetime import UTC, datetime
from io import BytesIO, StringIO
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from backend.config import Settings
from backend.ingestion import build_dedupe_key, load_ris_records_with_raw
from backend.models import (
    CriteriaResults,
    CriterionResult,
    HumanReviewUpdateRequest,
    ReviewCounts,
    ReviewItem,
    ReviewProjectDetail,
    ReviewProjectSummary,
    StudyRecord,
    _default_review_project_name,
    compute_consensus,
)


class ReviewProjectNotFoundError(FileNotFoundError):
    pass


class ReviewFileFormatError(ValueError):
    pass


class ReviewItemNotFoundError(FileNotFoundError):
    pass


class ReviewRisExportError(ValueError):
    pass


REVIEW_EXPORT_BASE_HEADERS = [
    "item_id",
    "source_id",
    "import_filename",
    "linked_run_id",
    "title",
    "abstract",
    "doi",
    "year",
    "journal",
    "source_files",
    "judgment_a",
    "reason_a",
    "judgment_b",
    "reason_b",
    "consensus",
    "human_judgment",
    "human_reason",
    "reviewed_at",
]


class ReviewManager:
    def __init__(self, settings: Settings) -> None:
        self._projects_dir = settings.review_sessions_dir.resolve()
        self._projects_dir.mkdir(parents=True, exist_ok=True)

    def list_projects(self) -> list[ReviewProjectSummary]:
        projects: list[ReviewProjectSummary] = []
        for path in sorted(self._projects_dir.iterdir(), reverse=True):
            if not path.is_dir():
                continue
            try:
                detail = self.get_project(path.name)
            except ReviewProjectNotFoundError:
                continue
            projects.append(
                ReviewProjectSummary(
                    project_id=detail.project_id,
                    name=detail.name,
                    source_filename=detail.source_filename,
                    source_filenames=detail.source_filenames,
                    created_at=detail.created_at,
                    updated_at=detail.updated_at,
                    counts=detail.counts,
                )
            )
        return sorted(projects, key=lambda item: item.updated_at, reverse=True)

    def create_project(self, name: str, uploads: list[tuple[str, bytes]]) -> ReviewProjectDetail:
        items, ris_records = self._load_items_and_ris_from_uploads(uploads)
        if ris_records:
            self._attach_ris_entries(items, ris_records)
        if not items:
            raise ReviewFileFormatError("No reviewable records were found in the uploaded files.")
        source_filenames = [filename for filename, _ in uploads]
        return self.create_project_from_items(
            name=name,
            source_filename=self._summarize_source_filenames(source_filenames),
            source_filenames=source_filenames,
            items=items,
        )

    def create_project_from_items(
        self,
        *,
        name: str,
        source_filename: str,
        source_filenames: list[str] | None = None,
        items: list[ReviewItem],
    ) -> ReviewProjectDetail:
        if not items:
            raise ReviewFileFormatError("No reviewable records were found in the provided results.")

        normalized_items = self._merge_items([], items)
        normalized_source_filenames = self._normalize_source_filenames(source_filenames or [source_filename])
        resolved_name = name.strip() or _default_review_project_name(normalized_source_filenames)
        project = ReviewProjectDetail(
            project_id=f"review_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}",
            name=resolved_name,
            source_filename=self._summarize_source_filenames(normalized_source_filenames),
            source_filenames=normalized_source_filenames,
            created_at=_utc_now(),
            updated_at=_utc_now(),
            counts=self._build_counts(normalized_items),
            items=normalized_items,
        )
        self._project_dir(project.project_id).mkdir(parents=True, exist_ok=False)
        self._write_project(project)
        return project

    def get_project(self, project_id: str) -> ReviewProjectDetail:
        path = self._project_path(project_id)
        if not path.exists():
            raise ReviewProjectNotFoundError(project_id)
        project = ReviewProjectDetail.model_validate_json(path.read_text(encoding="utf-8"))
        project.source_filenames = self._normalize_source_filenames(project.source_filenames or [project.source_filename])
        project.source_filename = self._summarize_source_filenames(project.source_filenames)
        project.items = self._merge_items([], project.items)
        project.counts = self._build_counts(project.items)
        return project

    def append_project_files(self, project_id: str, uploads: list[tuple[str, bytes]]) -> ReviewProjectDetail:
        project = self.get_project(project_id)
        items, ris_records = self._load_items_and_ris_from_uploads(uploads)
        if items and ris_records:
            self._attach_ris_entries(items, ris_records)
        if not items and not ris_records:
            raise ReviewFileFormatError("No reviewable records or RIS records were found in the uploaded files.")

        if items:
            project.items = self._merge_items(project.items, items)
        if ris_records:
            self._attach_ris_entries(project.items, ris_records)
        project.source_filenames = self._normalize_source_filenames(
            [*project.source_filenames, *(filename for filename, _ in uploads)]
        )
        project.source_filename = self._summarize_source_filenames(project.source_filenames)
        project.updated_at = _utc_now()
        project.counts = self._build_counts(project.items)
        self._write_project(project)
        return project

    def delete_project(self, project_id: str) -> None:
        directory = self._project_dir(project_id)
        if not directory.exists():
            raise ReviewProjectNotFoundError(project_id)
        import shutil

        shutil.rmtree(directory)

    def update_project_name(self, project_id: str, name: str) -> ReviewProjectDetail:
        project = self.get_project(project_id)
        project.name = name.strip()
        project.updated_at = _utc_now()
        self._write_project(project)
        return project

    def save_review(
        self,
        project_id: str,
        item_id: str,
        request: HumanReviewUpdateRequest,
    ) -> ReviewItem:
        project = self.get_project(project_id)
        target = next(
            (
                item
                for item in project.items
                if item.item_id == item_id or (item.item_id == item.source_id and item.source_id == item_id)
            ),
            None,
        )
        if target is None:
            raise ReviewItemNotFoundError(item_id)

        target.human_judgment = request.judgment
        target.human_reason = request.reason
        target.reviewed_at = _utc_now() if request.judgment or request.reason else None
        project.updated_at = _utc_now()
        project.counts = self._build_counts(project.items)
        self._write_project(project)
        return target

    def get_results_csv_path(self, project_id: str) -> Path:
        project = self.get_project(project_id)
        self._build_csv_export(project)
        return self._results_csv_path(project_id)

    def get_results_xlsx_path(self, project_id: str) -> Path:
        project = self.get_project(project_id)
        self._build_csv_export(project)
        self._build_xlsx_export(project_id)
        return self._results_xlsx_path(project_id)

    def get_included_ris_path(self, project_id: str) -> Path:
        project = self.get_project(project_id)
        included_items = [item for item in project.items if _final_status(item) == "included"]
        if not included_items:
            raise ReviewRisExportError("No records are currently marked as Included in RIS.")

        missing = [item for item in included_items if not (item.ris_entry or "").strip()]
        if missing:
            raise ReviewRisExportError(
                f"{len(missing)} included record(s) are missing original RIS entries. Attach the source RIS files before exporting."
            )

        entries = [_normalize_ris_entry(item.ris_entry or "") for item in included_items]
        path = self._included_ris_path(project_id)
        path.write_text("\n\n".join(entries) + "\n", encoding="utf-8")
        return path

    def _write_project(self, project: ReviewProjectDetail) -> None:
        project.source_filenames = self._normalize_source_filenames(project.source_filenames or [project.source_filename])
        project.source_filename = self._summarize_source_filenames(project.source_filenames)
        project.items = self._merge_items([], project.items)
        project.counts = self._build_counts(project.items)
        self._project_path(project.project_id).write_text(project.model_dump_json(indent=2), encoding="utf-8")

    def _load_items_from_uploads(self, uploads: list[tuple[str, bytes]]) -> list[ReviewItem]:
        items, ris_records = self._load_items_and_ris_from_uploads(uploads)
        if ris_records:
            self._attach_ris_entries(items, ris_records)
        return items

    def _load_items_and_ris_from_uploads(
        self,
        uploads: list[tuple[str, bytes]],
    ) -> tuple[list[ReviewItem], list[StudyRecord]]:
        items: list[ReviewItem] = []
        ris_records: list[StudyRecord] = []
        for filename, content in uploads:
            if filename.lower().endswith(".ris"):
                ris_records.extend(self._load_ris_records(filename, content))
                continue
            rows = self._load_rows(filename, content)
            items.extend(self._rows_to_items(rows, import_filename=filename))
        return items, ris_records

    def _load_rows(self, filename: str, content: bytes) -> list[dict[str, str]]:
        lower_name = filename.lower()
        if lower_name.endswith(".csv"):
            text = content.decode("utf-8-sig")
            return list(csv.DictReader(StringIO(text)))
        if lower_name.endswith(".xlsx"):
            workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(value or "").strip() for value in rows[0]]
            payload: list[dict[str, str]] = []
            for values in rows[1:]:
                payload.append(
                    {
                        headers[index]: "" if value is None else str(value)
                        for index, value in enumerate(values)
                        if index < len(headers) and headers[index]
                    }
                )
            return payload
        raise ReviewFileFormatError("Upload CSV/XLSX results files and optional RIS source files.")

    def _load_ris_records(self, filename: str, content: bytes) -> list[StudyRecord]:
        records: list[StudyRecord] = []
        for index, (record, raw_entry) in enumerate(load_ris_records_with_raw(content), start=1):
            title = _clean_value(record.get("title")) or _clean_value(record.get("primary_title"))
            raw_abstract = _clean_value(record.get("abstract")) or _clean_value(record.get("notes_abstract")) or ""
            if not title:
                continue
            doi = _normalize_doi(
                _clean_value(record.get("doi"))
                or _clean_value(record.get("first_doi"))
                or _clean_value(record.get("accession_number"))
            )
            year = _clean_value(record.get("year")) or _clean_value(record.get("publication_year"))
            journal = (
                _clean_value(record.get("journal_name"))
                or _clean_value(record.get("secondary_title"))
                or _clean_value(record.get("custom3"))
            )
            records.append(
                StudyRecord(
                    source_id=f"{filename}:{index}",
                    title=title,
                    abstract=raw_abstract or "No abstract available.",
                    doi=doi,
                    year=year,
                    journal=journal,
                    authors=[],
                    source_files=[filename],
                    dedupe_key=build_dedupe_key(title=title, doi=doi, abstract=raw_abstract),
                    ris_entry=raw_entry,
                )
            )
        return records

    def _attach_ris_entries(self, items: list[ReviewItem], ris_records: list[StudyRecord]) -> None:
        by_key: dict[str, StudyRecord] = {}
        by_title: dict[str, StudyRecord] = {}
        for record in ris_records:
            by_key[self._study_match_key(record)] = record
            title_key = _normalize_identity_token(record.title)
            if title_key and title_key not in by_title:
                by_title[title_key] = record

        for item in items:
            record = by_key.get(self._item_match_key(item))
            if record is None:
                record = by_title.get(_normalize_identity_token(item.title))
            if record is None:
                continue
            item.ris_entry = record.ris_entry
            item.ris_source_filename = record.source_files[0] if record.source_files else None

    def _rows_to_items(self, rows: list[dict[str, str]], *, import_filename: str) -> list[ReviewItem]:
        items: list[ReviewItem] = []
        for index, row in enumerate(rows, start=1):
            source_id = (row.get("source_id") or row.get("doi") or f"row-{index}").strip() or f"row-{index}"
            title = (
                (row.get("title") or "").strip()
                or (row.get("doi") or "").strip()
                or source_id
                or f"Untitled record {index}"
            )
            judgment_a = _normalize_optional_judgment(row.get("judgment_a"))
            judgment_b = _normalize_optional_judgment(row.get("judgment_b"))
            criterion_results_a = self._criterion_results_from_row(row, suffix="_a")
            criterion_results_b = self._criterion_results_from_row(row, suffix="_b")
            item = ReviewItem(
                item_id=(row.get("item_id") or "").strip() or source_id,
                sequence=index,
                source_id=source_id,
                title=title,
                abstract=(row.get("abstract") or "").strip(),
                doi=(row.get("doi") or "").strip() or None,
                year=(row.get("year") or "").strip() or None,
                journal=(row.get("journal") or "").strip() or None,
                source_files=_split_source_files((row.get("source_files") or "").strip()),
                import_filename=(row.get("import_filename") or "").strip() or import_filename,
                linked_run_id=(row.get("linked_run_id") or "").strip() or None,
                judgment_a=judgment_a,
                reason_a=(row.get("reason_a") or "").strip(),
                criterion_results_a=criterion_results_a,
                judgment_b=judgment_b,
                reason_b=(row.get("reason_b") or "").strip(),
                criterion_results_b=criterion_results_b,
                consensus=compute_consensus(judgment_a, judgment_b),
                human_judgment=_normalize_optional_judgment(row.get("human_judgment")),
                human_reason=(row.get("human_reason") or "").strip(),
                reviewed_at=(row.get("reviewed_at") or "").strip() or None,
            )
            items.append(item)
        return items

    def _merge_items(self, existing_items: list[ReviewItem], incoming_items: list[ReviewItem]) -> list[ReviewItem]:
        merged = [item.model_copy(deep=True) for item in existing_items]
        by_match_key = {self._item_match_key(item): item for item in merged}
        by_item_id = {item.item_id: item for item in merged}

        for incoming in incoming_items:
            candidate = incoming.model_copy(deep=True)
            if not candidate.item_id:
                candidate.item_id = candidate.source_id
            candidate.consensus = compute_consensus(candidate.judgment_a, candidate.judgment_b)

            existing = by_item_id.get(candidate.item_id) or by_match_key.get(self._item_match_key(candidate))
            if existing is None:
                if candidate.item_id in by_item_id:
                    candidate.item_id = f"{candidate.item_id}-{uuid4().hex[:8]}"
                merged.append(candidate)
                by_item_id[candidate.item_id] = candidate
                by_match_key[self._item_match_key(candidate)] = candidate
                continue

            existing.source_id = candidate.source_id
            existing.title = candidate.title
            existing.abstract = candidate.abstract
            existing.doi = candidate.doi
            existing.year = candidate.year
            existing.journal = candidate.journal
            existing.source_files = candidate.source_files
            existing.import_filename = candidate.import_filename or existing.import_filename
            existing.linked_run_id = candidate.linked_run_id or existing.linked_run_id
            existing.ris_entry = candidate.ris_entry or existing.ris_entry
            existing.ris_source_filename = candidate.ris_source_filename or existing.ris_source_filename
            existing.judgment_a = candidate.judgment_a
            existing.reason_a = candidate.reason_a
            existing.criterion_results_a = candidate.criterion_results_a
            existing.judgment_b = candidate.judgment_b
            existing.reason_b = candidate.reason_b
            existing.criterion_results_b = candidate.criterion_results_b
            existing.consensus = compute_consensus(existing.judgment_a, existing.judgment_b)
            if candidate.human_judgment is not None:
                existing.human_judgment = candidate.human_judgment
            if candidate.human_reason:
                existing.human_reason = candidate.human_reason
            if candidate.reviewed_at:
                existing.reviewed_at = candidate.reviewed_at

        for sequence, item in enumerate(merged, start=1):
            item.sequence = sequence

        return merged

    def _criterion_results_from_row(self, row: dict[str, str], *, suffix: str) -> CriteriaResults | None:
        # Suffix is "_a" or "_b". Look for inclusion_<token>_judgment_a etc.
        inclusion = self._read_header_criterion_group(row, "inclusion", suffix)
        exclusion = self._read_header_criterion_group(row, "exclusion", suffix)
        if not inclusion and not exclusion:
            return None
        return CriteriaResults(inclusion=inclusion, exclusion=exclusion)

    def _read_header_criterion_group(self, row: dict[str, str], group: str, suffix: str) -> list[CriterionResult]:
        grouped: dict[str, dict[str, str]] = {}
        prefix = f"{group}_"
        judgment_suffix = f"_judgment{suffix}"
        reason_suffix = f"_reason{suffix}"
        for key, value in row.items():
            if not key.startswith(prefix):
                continue
            if key.endswith(judgment_suffix):
                token = key[len(prefix) : -len(judgment_suffix)]
                grouped.setdefault(token, {})["judgment"] = str(value or "").strip()
            elif key.endswith(reason_suffix):
                token = key[len(prefix) : -len(reason_suffix)]
                grouped.setdefault(token, {})["reason"] = str(value or "").strip()

        items: list[CriterionResult] = []
        for token, fields in grouped.items():
            items.append(
                CriterionResult(
                    id=token or "criterion",
                    text=token or "Criterion",
                    judgment=_normalize_judgment(fields.get("judgment")),
                    reason=fields.get("reason") or "No reason available.",
                )
            )
        return items

    def _build_csv_export(self, project: ReviewProjectDetail) -> None:
        headers = self._export_headers(project)
        with self._results_csv_path(project.project_id).open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            for item in project.items:
                writer.writerow(self._export_row(project, item))

    def _build_xlsx_export(self, project_id: str) -> None:
        csv_path = self._results_csv_path(project_id)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Review Results"
        with csv_path.open("r", newline="", encoding="utf-8") as handle:
            reader = csv.reader(handle)
            for row in reader:
                worksheet.append(row)
        workbook.save(self._results_xlsx_path(project_id))

    def _export_headers(self, project: ReviewProjectDetail) -> list[str]:
        headers = list(REVIEW_EXPORT_BASE_HEADERS)
        for suffix in ("_a", "_b"):
            headers.extend(self._criterion_headers(project.items, "inclusion", suffix))
            headers.extend(self._criterion_headers(project.items, "exclusion", suffix))
        return headers

    def _export_row(self, project: ReviewProjectDetail, item: ReviewItem) -> list[str]:
        row = [
            item.item_id,
            item.source_id,
            item.import_filename or "",
            item.linked_run_id or "",
            item.title,
            item.abstract,
            item.doi or "",
            item.year or "",
            item.journal or "",
            " | ".join(item.source_files),
            item.judgment_a or "",
            item.reason_a,
            item.judgment_b or "",
            item.reason_b,
            item.consensus,
            item.human_judgment or "",
            item.human_reason,
            item.reviewed_at or "",
        ]
        for suffix, criteria_results in (
            ("_a", item.criterion_results_a),
            ("_b", item.criterion_results_b),
        ):
            row.extend(self._criterion_values(project.items, item, "inclusion", suffix, criteria_results))
            row.extend(self._criterion_values(project.items, item, "exclusion", suffix, criteria_results))
        return row

    def _criterion_headers(self, items: list[ReviewItem], group: str, suffix: str) -> list[str]:
        headers: list[str] = []
        for criterion_id in self._criterion_order(items, group):
            token = _safe_column_token(criterion_id)
            headers.append(f"{group}_{token}_judgment{suffix}")
            headers.append(f"{group}_{token}_reason{suffix}")
        return headers

    def _criterion_values(
        self,
        items: list[ReviewItem],
        item: ReviewItem,
        group: str,
        suffix: str,
        criteria_results: CriteriaResults | None,
    ) -> list[str]:
        if criteria_results is None:
            criterion_items: list[CriterionResult] = []
        else:
            criterion_items = (
                criteria_results.inclusion if group == "inclusion" else criteria_results.exclusion
            )
        criterion_map = {criterion.id: criterion for criterion in criterion_items}
        values: list[str] = []
        for criterion_id in self._criterion_order(items, group):
            criterion = criterion_map.get(criterion_id)
            values.append(criterion.judgment if criterion else "")
            values.append(criterion.reason if criterion else "")
        return values

    def _criterion_order(self, items: list[ReviewItem], group: str) -> list[str]:
        ordered: list[str] = []
        seen: set[str] = set()
        for item in items:
            for criteria_results in (item.criterion_results_a, item.criterion_results_b):
                if criteria_results is None:
                    continue
                criterion_items = (
                    criteria_results.inclusion if group == "inclusion" else criteria_results.exclusion
                )
                for criterion in criterion_items:
                    if criterion.id in seen:
                        continue
                    seen.add(criterion.id)
                    ordered.append(criterion.id)
        return ordered

    def _build_counts(self, items: list[ReviewItem]) -> ReviewCounts:
        needs_review_items = [item for item in items if _requires_human_review(item)]
        return ReviewCounts(
            total=len(items),
            consensus_agree=sum(1 for item in items if item.consensus == "agree"),
            consensus_partial=sum(1 for item in items if item.consensus == "partial"),
            consensus_conflict=sum(1 for item in items if item.consensus == "conflict"),
            consensus_pending=sum(1 for item in items if item.consensus == "pending"),
            reviewed=sum(1 for item in items if item.human_judgment is not None),
            needs_review_total=len(needs_review_items),
            needs_review_reviewed=sum(1 for item in needs_review_items if item.human_judgment is not None),
            human_yes=sum(1 for item in items if item.human_judgment == "yes"),
            human_no=sum(1 for item in items if item.human_judgment == "no"),
            human_maybe=sum(1 for item in items if item.human_judgment == "maybe"),
        )

    def delete_project(self, project_id: str) -> None:
        directory = self._project_dir(project_id)
        if not directory.exists() or not directory.is_dir():
            raise ReviewProjectNotFoundError(project_id)
        import shutil

        shutil.rmtree(directory)

    def _project_dir(self, project_id: str) -> Path:
        return self._projects_dir / project_id

    def _project_path(self, project_id: str) -> Path:
        return self._project_dir(project_id) / "project.json"

    def _results_csv_path(self, project_id: str) -> Path:
        return self._project_dir(project_id) / "results.csv"

    def _results_xlsx_path(self, project_id: str) -> Path:
        return self._project_dir(project_id) / "results.xlsx"

    def _included_ris_path(self, project_id: str) -> Path:
        return self._project_dir(project_id) / "included.ris"

    def _item_match_key(self, item: ReviewItem) -> str:
        doi = _normalize_identity_token(item.doi)
        if doi:
            return f"doi:{doi}"

        title = _normalize_identity_token(item.title)
        year = _normalize_identity_token(item.year)
        journal = _normalize_identity_token(item.journal)
        if title:
            return f"title:{title}|year:{year}|journal:{journal}"

        source_id = _normalize_identity_token(item.source_id)
        import_filename = _normalize_identity_token(item.import_filename)
        return f"source:{source_id}|file:{import_filename}"

    def _study_match_key(self, study: StudyRecord) -> str:
        doi = _normalize_identity_token(study.doi)
        if doi:
            return f"doi:{doi}"

        title = _normalize_identity_token(study.title)
        year = _normalize_identity_token(study.year)
        journal = _normalize_identity_token(study.journal)
        if title:
            return f"title:{title}|year:{year}|journal:{journal}"

        source_id = _normalize_identity_token(study.source_id)
        return f"source:{source_id}|file:"

    def _normalize_source_filenames(self, source_filenames: list[str]) -> list[str]:
        normalized: list[str] = []
        for value in source_filenames:
            token = str(value or "").strip()
            if token and token not in normalized:
                normalized.append(token)
        return normalized

    def _summarize_source_filenames(self, source_filenames: list[str]) -> str:
        normalized = self._normalize_source_filenames(source_filenames)
        if not normalized:
            return "Imported results"
        if len(normalized) == 1:
            return normalized[0]
        return f"{normalized[0]} + {len(normalized) - 1} more"


def _utc_now() -> str:
    return datetime.now(UTC).isoformat()


def _normalize_judgment(value: object) -> str:
    token = str(value or "").strip().lower()
    if token in {"yes", "no", "maybe"}:
        return token
    return "maybe"


def _normalize_optional_judgment(value: object) -> str | None:
    token = str(value or "").strip().lower()
    if token in {"yes", "no", "maybe"}:
        return token
    return None


def _split_source_files(value: str) -> list[str]:
    return [part.strip() for part in value.split("|") if part.strip()]


def _safe_column_token(value: str) -> str:
    token = "".join(char if char.isalnum() else "_" for char in value.strip())
    token = token.strip("_")
    return token or "criterion"


def _normalize_identity_token(value: str | None) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _final_status(item: ReviewItem) -> str:
    if item.human_judgment == "yes":
        return "included"
    if item.human_judgment == "no":
        return "excluded"
    if item.human_judgment == "maybe":
        return "needs_review"
    if item.judgment_a == "yes" and item.judgment_b == "yes":
        return "included"
    if item.judgment_a == "no" and item.judgment_b == "no":
        return "excluded"
    return "needs_review"


def _requires_human_review(item: ReviewItem) -> bool:
    return not (
        (item.judgment_a == "yes" and item.judgment_b == "yes")
        or (item.judgment_a == "no" and item.judgment_b == "no")
    )


def _normalize_ris_entry(value: str) -> str:
    cleaned = value.strip()
    if not cleaned:
        return cleaned
    if "ER  -" not in cleaned:
        cleaned = f"{cleaned}\nER  -"
    return cleaned


def _clean_value(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, list):
        joined = " ".join(str(item).strip() for item in value if str(item).strip())
        return joined or None
    cleaned = str(value).strip()
    return cleaned or None


def _normalize_doi(value: str | None) -> str | None:
    if not value:
        return None
    cleaned = value.strip().lower()
    cleaned = cleaned.removeprefix("https://doi.org/")
    cleaned = cleaned.removeprefix("http://doi.org/")
    cleaned = cleaned.removeprefix("doi:")
    return cleaned.strip() or None
